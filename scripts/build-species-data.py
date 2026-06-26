#!/usr/bin/env python3
"""Regenerate constants/iucnStatus.ts and constants/avonet.ts (+ recovery audit).

Both files map IOC v15.2 common names to licensed open data, joined through the
scientific names in constants/birdLatin.ts. We ship static maps (no runtime API
calls) so the Guide tab works offline and stays commercially clean.

Two-stage join per dataset:
  1. Direct match — the IOC binomial equals a source binomial.
  2. Recovery — for the misses (IOC genus renames / splits the source files under
     a synonym), recover via:
       - genus-synonym: same species epithet, exactly one candidate in the source
         (catches genus reassignments, e.g. Icthyophaga vocifer -> Haliaeetus vocifer).
       - crosswalk (AVONET only): the IOC binomial appears in AVONET's
         BirdLife<->eBird / BirdLife<->BirdTree crosswalk; use the partner row.
     Every recovered entry is LOWER confidence than a direct match and is written
     to scripts/species-data-recovered.csv for manual review.

INPUTS (fetch first, then pass their paths):

1. IUCN status (Wikidata P141, CC0). Wikidata's status entities no longer carry
   P528 IUCN codes, so we pull the raw status Q-id per species and map it here.
   Run at https://query.wikidata.org/ (or curl the /sparql endpoint, Accept:
   application/sparql-results+json) and save JSON:

       SELECT ?sci ?status WHERE {
         ?species wdt:P141 ?status; wdt:P171* wd:Q5113; wdt:P225 ?sci.
       }

2. AVONET (Tobias et al. 2022, Ecology Letters, CC BY 4.0):
       https://ndownloader.figshare.com/files/34480856   (AVONET ...dataset 1.xlsx)

USAGE:
   python3 scripts/build-species-data.py <wikidata.json> <avonet.xlsx>

Refresh cadence: IUCN ~once a year; AVONET is static (one published release).
"""
import sys, json, re, os, csv
from collections import defaultdict

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_OUT = os.path.join(ROOT, 'scripts', 'species-data-recovered.csv')

# Wikidata status entity -> IUCN code (verified 2026-06-26 by label lookup).
QCODE = {
    'Q211005': 'LC', 'Q719675': 'NT', 'Q278113': 'VU', 'Q96377276': 'EN',
    'Q219127': 'CR', 'Q237350': 'EX', 'Q239509': 'EW', 'Q3245245': 'DD',
}
# When Wikidata lists more than one assessment, keep the most-threatened.
PRIORITY = {'CR': 0, 'EN': 1, 'VU': 2, 'NT': 3, 'LC': 4, 'EW': 5, 'EX': 6, 'DD': 7}
MIG = {1: 'sedentary', 2: 'partial', 3: 'migratory'}

# audit rows collected across both datasets: dict per recovered entry
AUDIT = []


def norm(s):
    """Lowercased genus + species (drop any subspecies/qualifier tokens)."""
    return ' '.join((s or '').strip().lower().split()[:2])


def load_bird_latin():
    t = open(os.path.join(ROOT, 'constants', 'birdLatin.ts')).read()
    return eval(re.search(r'BIRD_LATIN[^=]*=\s*(\{.*?\n\})', t, re.S).group(1))  # noqa: S307


def ioc_meta():
    """Return (order list, name->family)."""
    t = open(os.path.join(ROOT, 'constants', 'birdNames.ts')).read()
    order, fam, cur = [], {}, None
    for m in re.finditer(r'family:\s*"((?:[^"\\]|\\.)*)"|name:\s*"((?:[^"\\]|\\.)*)"', t):
        if m.group(1) is not None:
            cur = m.group(1)
        else:
            order.append(m.group(2)); fam[m.group(2)] = cur
    return order, fam


def genus_families(latin, fam):
    """IOC genus (lowercase) -> set of IOC families that use it. Lets us reject a
    same-epithet 'genus rename' whose genus belongs to a different family (e.g.
    Red Grouse 'scotica' must not match the crossbill Loxia scotica)."""
    gf = defaultdict(set)
    for common, lat in latin.items():
        g = norm(lat).split(' ')[0]
        if g and common in fam:
            gf[g].add(fam[common])
    return gf


def sci_case(n):
    """'loxia scotica' -> 'Loxia scotica' (proper binomial casing)."""
    parts = n.split()
    return ' '.join([parts[0].capitalize()] + parts[1:]) if parts else n


def ordered_keys(d, order):
    oset = set(order)
    return [n for n in order if n in d] + [k for k in d if k not in oset]


def epithet_index(norm_names):
    idx = defaultdict(set)
    for nm in norm_names:
        parts = nm.split()
        if len(parts) == 2:
            idx[parts[1]].add(nm)
    return idx


def genus_synonym(latin, src_index, target_family, genus_fams):
    """If exactly one source binomial shares this epithet AND its genus is one
    IOC uses in the same family, return it (a likely genus rename). The family
    guard rejects coincidental cross-family epithet collisions."""
    n = norm(latin)
    parts = n.split()
    if len(parts) != 2 or len(parts[1]) < 4:
        return None
    cands = {c for c in src_index.get(parts[1], set()) if c != n}
    if len(cands) != 1:
        return None
    match = next(iter(cands))
    cand_genus = match.split(' ')[0]
    if target_family not in genus_fams.get(cand_genus, ()):
        return None
    return match


# ─── IUCN ────────────────────────────────────────────────────────────────────
def build_iucn(wikidata_json, latin, order, fam, genus_fams):
    raw = json.load(open(wikidata_json))
    by_sci = {}
    for r in raw['results']['bindings']:
        sci = norm(r.get('sci', {}).get('value', ''))
        q = r.get('status', {}).get('value', '').split('/')[-1]
        code = QCODE.get(q)
        if not sci or not code:
            continue
        if sci not in by_sci or PRIORITY[code] < PRIORITY[by_sci[sci]]:
            by_sci[sci] = code

    iucn, resolved = {}, {}
    for common, lat in latin.items():
        code = by_sci.get(norm(lat))
        if code:
            iucn[common] = code

    # Recovery pass for the misses.
    idx = epithet_index(by_sci.keys())
    recovered = 0
    for common, lat in latin.items():
        if common in iucn or not lat:
            continue
        match = genus_synonym(lat, idx, fam.get(common), genus_fams)
        if match:
            iucn[common] = by_sci[match]
            resolved[common] = match
            recovered += 1
            AUDIT.append({'dataset': 'IUCN', 'ioc_name': common, 'ioc_latin': lat,
                          'method': 'genus-synonym', 'confidence': 'medium',
                          'matched_name': sci_case(match), 'value': by_sci[match], 'note': ''})

    write_iucn_ts(iucn, order, recovered)
    return resolved


def write_iucn_ts(iucn, order, recovered):
    keys = ordered_keys(iucn, order)
    out = os.path.join(ROOT, 'constants', 'iucnStatus.ts')
    with open(out, 'w') as f:
        f.write(
            "// Generated by scripts/build-species-data.py - do not edit by hand.\n"
            "// Source: Wikidata P141 (IUCN conservation status), CC0, joined to IOC\n"
            "// v15.2 common names via constants/birdLatin.ts (scientific-name match).\n"
            "// Status itself originates with the IUCN Red List of Threatened Species.\n"
            f"// Includes {recovered} entries recovered via genus-synonym matching -\n"
            "// see scripts/species-data-recovered.csv. Unmatched names resolve to 'NE'.\n\n"
            "import { IUCNStatusCode } from './Colors';\n\n"
            "const IUCN_STATUS: Record<string, IUCNStatusCode> = {\n")
        for n in keys:
            f.write(f"  {json.dumps(n)}: '{iucn[n]}',\n")
        f.write(
            "};\n\n"
            "// Conservation status for a species by its IOC common name. Unknown or\n"
            "// unassessed species return 'NE' (Not Evaluated).\n"
            "export function statusFor(name: string): IUCNStatusCode {\n"
            "  return IUCN_STATUS[name] ?? 'NE';\n"
            "}\n\n"
            "export default IUCN_STATUS;\n")
    print(f"wrote {out}: {len(keys)} entries ({recovered} recovered)")


# ─── AVONET ──────────────────────────────────────────────────────────────────
def harvest_sheet(wb, sheet, sci_col):
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    hdr = list(next(it))
    idx = {h: i for i, h in enumerate(hdr)}

    def num(row, c):
        v = row[idx[c]]
        return round(float(v), 1) if isinstance(v, (int, float)) else None

    def clean(v):
        return None if v in (None, 'NA', '', 'na') else v

    def mig(v):
        try:
            return MIG.get(int(float(v)))
        except (TypeError, ValueError):
            return None

    out = {}
    for row in it:
        sci = norm(row[idx[sci_col]] or '')
        if not sci:
            continue
        out[sci] = {
            'mass': num(row, 'Mass'), 'wingChord': num(row, 'Wing.Length'),
            'habitat': clean(row[idx['Habitat']]),
            'diet': clean(row[idx['Trophic.Niche']]),
            'migration': mig(row[idx['Migration']]),
        }
    return out


def load_crosswalks(wb):
    """norm(binomial) -> list of (partner norm binomial, match_type), both directions."""
    xwalk = defaultdict(list)
    for sheet, a, b in (('BirdLife-eBird crosswalk', 'Species1', 'Species2'),
                        ('BirdLife–BirdTree crosswalk', 'Species1', 'Species3')):
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        hdr = list(next(it))
        ia, ib = hdr.index(a), hdr.index(b)
        it_mt = hdr.index('Match.type') if 'Match.type' in hdr else None
        for row in it:
            na, nb = norm(row[ia] or ''), norm(row[ib] or '')
            mt = (row[it_mt] if it_mt is not None else '') or ''
            if na and nb and na != nb:
                xwalk[na].append((nb, mt))
                xwalk[nb].append((na, mt))
    return xwalk


def build_avonet(xlsx, latin, order, iucn_resolved, fam, genus_fams):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    av = harvest_sheet(wb, 'AVONET1_BirdLife', 'Species1')
    for sheet, col in (('AVONET2_eBird', 'Species2'), ('AVONET3_BirdTree', 'Species3')):
        for k, v in harvest_sheet(wb, sheet, col).items():
            av.setdefault(k, v)
    xwalk = load_crosswalks(wb)
    idx = epithet_index(av.keys())

    ioc = {}
    for common, lat in latin.items():
        rec = av.get(norm(lat))
        if rec:
            ioc[common] = rec

    recovered = 0
    for common, lat in latin.items():
        if common in ioc or not lat:
            continue
        match, method, conf = None, None, None
        # 1:1 crosswalk first (a named taxonomic equivalent), then genus-synonym,
        # then any other crosswalk partner (splits/lumps — lowest confidence).
        partners = [(p, mt) for p, mt in xwalk.get(norm(lat), []) if p in av]
        one_to_one = [(p, mt) for p, mt in partners if re.search(r'\b1\b.*\b1\b', mt)]
        gs = genus_synonym(lat, idx, fam.get(common), genus_fams)
        if one_to_one:
            match, mt = one_to_one[0]; method, conf = f'crosswalk ({mt})', 'medium'
        elif gs:
            match, method, conf = gs, 'genus-synonym', 'medium'
        elif partners:
            match, mt = partners[0]; method, conf = f'crosswalk ({mt})', 'low'
        if not match:
            continue
        ioc[common] = av[match]
        recovered += 1
        r = av[match]
        val = f"{r['mass']}g/{r['wingChord']}mm/{r['habitat']}/{r['diet']}/{r['migration']}"
        note = 'corroborated by IUCN' if iucn_resolved.get(common) == match else ''
        AUDIT.append({'dataset': 'AVONET', 'ioc_name': common, 'ioc_latin': lat,
                      'method': method, 'confidence': conf,
                      'matched_name': sci_case(match), 'value': val, 'note': note})

    write_avonet_ts(ioc, order, recovered)


def write_avonet_ts(ioc, order, recovered):
    keys = ordered_keys(ioc, order)

    def fmt(v):
        if v is None:
            return 'null'
        if isinstance(v, (int, float)):
            return repr(v)
        return json.dumps(v)

    out = os.path.join(ROOT, 'constants', 'avonet.ts')
    with open(out, 'w') as f:
        f.write(
            "// Generated by scripts/build-species-data.py - do not edit by hand.\n"
            "// Source: AVONET (Tobias et al. 2022, Ecology Letters), CC BY 4.0,\n"
            "// figshare 10.6084/m9.figshare.16586228. BirdLife/eBird/BirdTree sheets\n"
            "// merged, joined to IOC v15.2 common names via constants/birdLatin.ts.\n"
            f"// Includes {recovered} entries recovered via genus-synonym / crosswalk -\n"
            "// see scripts/species-data-recovered.csv. Attribution is mandatory (CC BY).\n\n"
            "export type Migration = 'sedentary' | 'partial' | 'migratory';\n"
            "export interface AvonetRecord {\n"
            "  mass: number | null;       // body mass, grams\n"
            "  wingChord: number | null;  // wing length, mm\n"
            "  habitat: string | null;\n"
            "  diet: string | null;       // AVONET trophic niche\n"
            "  migration: Migration | null;\n"
            "}\n\n"
            "const AVONET: Record<string, AvonetRecord> = {\n")
        for n in keys:
            r = ioc[n]
            f.write(
                f"  {json.dumps(n)}: {{ mass: {fmt(r['mass'])}, wingChord: {fmt(r['wingChord'])}, "
                f"habitat: {fmt(r['habitat'])}, diet: {fmt(r['diet'])}, migration: {fmt(r['migration'])} }},\n")
        f.write(
            "};\n\n"
            "// AVONET morphology + ecology for a species by its IOC common name.\n"
            "// Returns null when the species isn't in AVONET (drives the '-' UI state).\n"
            "export function avonetFor(name: string): AvonetRecord | null {\n"
            "  return AVONET[name] ?? null;\n"
            "}\n\n"
            "export default AVONET;\n")
    print(f"wrote {out}: {len(keys)} entries ({recovered} recovered)")


def write_audit_csv(order):
    rank = {n: i for i, n in enumerate(order)}
    AUDIT.sort(key=lambda r: (r['dataset'], rank.get(r['ioc_name'], 1e9)))
    with open(CSV_OUT, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['dataset', 'ioc_name', 'ioc_latin',
                                          'method', 'confidence', 'matched_name', 'value', 'note'])
        w.writeheader()
        w.writerows(AUDIT)
    print(f"wrote {CSV_OUT}: {len(AUDIT)} recovered entries for review")


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    latin = load_bird_latin()
    order, fam = ioc_meta()
    gfam = genus_families(latin, fam)
    iucn_resolved = build_iucn(sys.argv[1], latin, order, fam, gfam)
    build_avonet(sys.argv[2], latin, order, iucn_resolved, fam, gfam)
    write_audit_csv(order)


if __name__ == '__main__':
    main()
